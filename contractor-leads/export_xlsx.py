"""
Export enriched_leads.json to a formatted xlsx the user can work from.

Columns (in the order the user asked for — best social first, then fallbacks):
    Company Name | Address | County | Category | Phone | Primary Contact |
    Facebook | Instagram | LinkedIn | Other Socials | Website | Email(s) |
    Google Maps | Rating | Reviews | Status | Enrichment Notes

'Primary Contact' is the best lead channel per the user's priority rules:
    facebook -> instagram -> linkedin -> other social -> website -> email

Usage:
    python export_xlsx.py --input enriched_leads.json --output leads.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = [
    "Company Name", "Address", "County", "Category", "Phone",
    "Primary Contact", "Facebook", "Instagram", "LinkedIn",
    "Other Socials", "Website", "Email(s)", "Google Maps",
    "Rating", "Reviews", "Status", "Outreach Sent?", "Response?",
    "Notes", "Enrichment Notes",
]

COLUMN_WIDTHS = {
    "Company Name": 32, "Address": 40, "County": 18, "Category": 24,
    "Phone": 16, "Primary Contact": 45, "Facebook": 40, "Instagram": 32,
    "LinkedIn": 32, "Other Socials": 32, "Website": 32, "Email(s)": 38,
    "Google Maps": 32, "Rating": 8, "Reviews": 10, "Status": 14,
    "Outreach Sent?": 14, "Response?": 12, "Notes": 40, "Enrichment Notes": 28,
}


def pick_primary_contact(lead: dict[str, Any]) -> str:
    for key in ("facebook", "instagram", "linkedin", "twitter", "youtube", "tiktok"):
        if lead.get(key):
            return lead[key]
    if lead.get("website"):
        return lead["website"]
    if lead.get("emails"):
        return "mailto:" + lead["emails"][0]
    return ""


def other_socials(lead: dict[str, Any]) -> str:
    extras = []
    for key in ("twitter", "youtube", "tiktok"):
        if lead.get(key):
            extras.append(f"{key}: {lead[key]}")
    return "\n".join(extras)


def build_workbook(leads: list[dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header row
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Sort: leads with FB first, then with any social, then rest
    def sort_key(l: dict[str, Any]) -> tuple:
        has_fb = 0 if l.get("facebook") else 1
        has_social = 0 if any(l.get(k) for k in ("facebook", "instagram", "linkedin")) else 1
        has_contact = 0 if (l.get("phone") or l.get("website") or l.get("emails")) else 1
        return (has_fb, has_social, has_contact, -(l.get("review_count") or 0))

    leads_sorted = sorted(leads, key=sort_key)

    for r, lead in enumerate(leads_sorted, start=2):
        emails = lead.get("emails") or []
        row = [
            lead.get("name", ""),
            lead.get("address", ""),
            lead.get("county", ""),
            lead.get("category_query", ""),
            lead.get("phone", ""),
            pick_primary_contact(lead),
            lead.get("facebook", ""),
            lead.get("instagram", ""),
            lead.get("linkedin", ""),
            other_socials(lead),
            lead.get("website", ""),
            "\n".join(emails),
            lead.get("google_maps_url", ""),
            lead.get("rating", ""),
            lead.get("review_count", ""),
            lead.get("business_status", "").replace("OPERATIONAL", "Open"),
            "",  # Outreach Sent? - user fills in
            "",  # Response?
            "",  # Notes
            lead.get("enrichment_notes", ""),
        ]
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    # Column widths
    for idx, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(h, 18)

    # Freeze header row + first column
    ws.freeze_panes = "B2"

    # Auto filter + convert to table for easy sorting
    last_col = get_column_letter(len(HEADERS))
    last_row = len(leads_sorted) + 1
    if last_row > 1:
        table_ref = f"A1:{last_col}{last_row}"
        tbl = Table(displayName="Leads", ref=table_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(tbl)

    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Central NJ Contractor Leads — Summary"
    summary["A1"].font = Font(name="Arial", bold=True, size=14)
    summary["A3"] = "Total leads"
    summary["B3"] = len(leads_sorted)
    summary["A4"] = "With Facebook"
    summary["B4"] = f'=COUNTIF(Leads!G:G,"<>")-1'
    summary["A5"] = "With Instagram"
    summary["B5"] = f'=COUNTIF(Leads!H:H,"<>")-1'
    summary["A6"] = "With LinkedIn"
    summary["B6"] = f'=COUNTIF(Leads!I:I,"<>")-1'
    summary["A7"] = "With Website"
    summary["B7"] = f'=COUNTIF(Leads!K:K,"<>")-1'
    summary["A8"] = "With Email"
    summary["B8"] = f'=COUNTIF(Leads!L:L,"<>")-1'
    summary["A10"] = "By County"
    summary["A10"].font = Font(name="Arial", bold=True)
    counties = {}
    for l in leads_sorted:
        counties[l.get("county", "Unknown")] = counties.get(l.get("county", "Unknown"), 0) + 1
    for i, (county, n) in enumerate(sorted(counties.items()), start=11):
        summary.cell(row=i, column=1, value=county)
        summary.cell(row=i, column=2, value=n)

    for col in "AB":
        summary.column_dimensions[col].width = 26

    return wb


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input", default="enriched_leads.json")
    ap.add_argument("--output", default="leads.xlsx")
    args = ap.parse_args()

    with open(args.input) as f:
        leads = json.load(f)

    wb = build_workbook(leads)
    wb.save(args.output)
    print(f"Wrote {len(leads)} leads -> {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
